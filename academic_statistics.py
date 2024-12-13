import openpyxl
import argparse
import statistics
import logging as log

log.basicConfig(level=log.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


def read_xlsx(file_path: str) -> list[dict]:
    """
    Чтение данных из XLSX файла.

    Args:
        file_path (str): Путь к входному файлу.

    Returns:
        list[dict]: Список студентов с их данными о посещаемости.
    """
    log.info("Чтение данных из файла %s.", file_path)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    students = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        attendance = list(row[1:])
        students.append({'name': name, 'attendance': attendance})

    log.info("Данные успешно считаны для %d студентов.", len(students))
    return students


def validate_attendance(attendance: list) -> list:
    """
    Очистка записей о посещаемости.

    Args:
        attendance (list): Список записей о посещаемости.

    Returns:
        list: Очищенные записи о посещаемости.
    """
    return [str(record) if isinstance(record, int) else record for record in attendance]


def get_avg_grade(attendance: list[str]) -> int:
    """
    Вычисление среднего балла.

    Args:
        attendance (list[str]): Список записей о посещаемости.

    Returns:
        int: Средний балл.
    """
    # Преобразование каждого элемент в строку, чтобы использовать isdigit
    grades = [int(record) for record in attendance if str(record).isdigit()]
    return round(statistics.mean(grades)) if grades else 0


def check_auto_pass(attendance: list[str], required_percentage: float) -> str:
    """
    Проверка, получил ли студент автомат.

    Args:
        attendance (list[str]): Список записей о посещаемости.
        required_percentage (float): Необходимый процент посещаемости.

    Returns:
        str: 'Да', если студент получил автомат, иначе 'Нет'.
    """
    total_days = len(attendance)
    # Приводим все записи к строкам, чтобы избежать ошибок
    attended_days = sum(1 for record in attendance if str(record) == '+' or str(record).isdigit())
    attendance_percentage = (attended_days / total_days) * 100 if total_days > 0 else 0
    return 'Да' if attendance_percentage >= required_percentage else 'Нет'


def count_absences(attendance: list[str]) -> int:
    """
    Подсчёт пропусков.

    Args:
        attendance (list[str]): Список записей о посещаемости.

    Returns:
        int: Количество пропусков.
    """
    return attendance.count('-')


def calculate_statistics(students: list[dict]) -> dict:
    """
    Общая статистика по курсу.

    Args:
        students (list[dict]): Список студентов с их данными о посещаемости.

    Returns:
        dict: Средний балл и процент посещаемости.
    """
    all_grades = []
    total_attended = 0
    total_days = 0

    for student in students:
        grades = [int(record) for record in student['attendance'] if isinstance(record, (int, str)) and str(record).isdigit()]
        all_grades.extend(grades)
        total_days += len(student['attendance'])
        total_attended += sum(1 for record in student['attendance'] if isinstance(record, (int, str)) and (record == '+' or str(record).isdigit()))

    avg_course_grade = round(statistics.mean(all_grades)) if all_grades else 0
    attendance_percentage = round((total_attended / total_days) * 100, 2) if total_days else 0

    return {'avg_course_grade': avg_course_grade, 'attendance_percentage': attendance_percentage}


def get_top_students(students: list[dict], top_n: int = 3) -> list[str]:
    """
    Определение топ-студентов по средней оценке.

    Args:
        students (list[dict]): Список студентов с их данными о посещаемости.
        top_n (int): Количество лучших студентов.

    Returns:
        list: Список лучших студентов.
    """
    students_with_grades = [(student['name'], get_avg_grade(student['attendance'])) for student in students]
    sorted_students = sorted(students_with_grades, key=lambda x: x[1], reverse=True)
    return sorted_students[:top_n]


def get_bad_students(students: list[dict], max_absences: int = 3) -> list[str]:
    """
    Определение студентов с плохой посещаемостью.

    Args:
        students (list[dict]): Список студентов с их данными о посещаемости.
        max_absences (int): Максимальное количество пропусков.

    Returns:
        list: Список студентов с плохой посещаемостью.
    """
    return [student['name'] for student in students if count_absences(student['attendance']) > max_absences]


def count_students_with_auto_pass(students: list[dict], threshold: float) -> tuple[int, int]:
    """
    Подсчёт студентов с автоматом и без автомата.

    Args:
        students (list[dict]): Список студентов с их данными о посещаемости.
        threshold (float): Процент посещаемости для автомата.

    Returns:
        tuple: Количество студентов с автоматом и без.
    """
    passed = 0
    failed = 0
    for student in students:
        if check_auto_pass(student['attendance'], threshold) == 'Да':
            passed += 1
        else:
            failed += 1
    return passed, failed


def save_results(workbook, results: list[dict]):
    """
    Сохранение результатов анализа.

    Args:
        workbook: Объект Workbook.
        results (list[dict]): Результаты анализа по студентам.
    """
    sheet_results = workbook.active
    sheet_results.title = "Результаты"
    sheet_results.append(["ФИО", "Автомат", "Средняя оценка", "Пропуски"])

    for result in results:
        sheet_results.append([result['name'], result['auto_pass'], result['avg_grade'], result['absences']])


def save_statistics(workbook, stats: dict):
    """
    Сохранение общей статистики.

    Args:
        workbook: Объект Workbook.
        stats (dict): Статистика курса.
    """
    sheet_summary = workbook.create_sheet("Статистика")
    sheet_summary.append(["Средняя оценка по курсу", "Процент посещаемости"])
    sheet_summary.append([stats['avg_course_grade'], stats['attendance_percentage']])


def save_top_students(workbook, top_students: list):
    """
    Сохранение списка топ студентов.

    Args:
        workbook: Объект Workbook.
        top_students (list): Список топ студентов.
    """
    sheet_top = workbook.create_sheet("Топ-студенты")
    sheet_top.append(["ФИО топ-студента", "Средняя оценка"])
    for name, grade in top_students:
        sheet_top.append([name, grade])


def save_bad_students(workbook, bad_students: list):
    """
    Сохранение списка 'плохих' студентов.

    Args:
        workbook: Объект Workbook.
        bad_students (list): Список проблемных студентов.
    """
    sheet_problems = workbook.create_sheet("Проблемные студенты")
    sheet_problems.append(["ФИО 'проблемного' студента"])
    for name in bad_students:
        sheet_problems.append([name])


def save_auto_pass_statistics(workbook, auto_pass_stats: tuple):
    """
    Сохранение статистики по автоматам.

    Args:
        workbook: Объект Workbook.
        auto_pass_stats (tuple): Кортеж с количеством студентов с автоматом и без.
    """
    sheet_auto_pass = workbook.create_sheet("Автоматы")
    sheet_auto_pass.append(["С автоматом", "Без автомата"])
    sheet_auto_pass.append([auto_pass_stats[0], auto_pass_stats[1]])


def analyze_and_save(input_file: str, output_file: str, threshold: float):
    """
    Анализ оценок и посещаемости, сохранение результатов.

    Args:
        input_file (str): Путь к входному файлу XLSX.
        output_file (str): Путь к выходному файлу XLSX.
        threshold (float): Процент посещаемости для автомата.
    """
    students = read_xlsx(input_file)
    workbook = openpyxl.Workbook()

    # Результаты анализа по каждому студенту
    results = []
    for student in students:
        attendance = validate_attendance(student['attendance'])
        avg_grade = get_avg_grade(attendance)
        auto_pass = check_auto_pass(attendance, threshold)
        absences = count_absences(attendance)

        results.append({
            'name': student['name'],
            'auto_pass': auto_pass,
            'avg_grade': avg_grade,
            'absences': absences
        })

    save_results(workbook, results)

    # Общая статистика курса
    stats = calculate_statistics(students)
    save_statistics(workbook, stats)

    # Топ студенты
    top_students = get_top_students(students)
    save_top_students(workbook, top_students)

    # Проблемные студенты
    bad_students = get_bad_students(students)
    save_bad_students(workbook, bad_students)

    # Статистика по автоматам
    auto_pass_stats = count_students_with_auto_pass(students, threshold)
    save_auto_pass_statistics(workbook, auto_pass_stats)

    # Сохранение файла
    workbook.save(output_file)
    log.info(f"Результаты сохранены в {output_file}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Анализ посещаемости и оценок студентов.")
    parser.add_argument("input_file", help="Путь к входному файлу XLSX")
    parser.add_argument("output_file", help="Путь к выходному файлу XLSX")
    parser.add_argument("--threshold", type=float, default=75, help="Процент посещаемости для автомата")

    args = parser.parse_args()
    analyze_and_save(args.input_file, args.output_file, args.threshold)
